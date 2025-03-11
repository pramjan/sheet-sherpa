# excel_analyzer.py

import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from io import BytesIO
import time
import altair as alt
import re

# Set the page layout to wide and a more appealing theme
st.set_page_config(layout="wide", page_title="Excel Analyzer", page_icon=":bar_chart:")

# Custom CSS for improved aesthetics
st.markdown(
    """
<style>
    /* Reduce padding and margins */
    .reportview-container .main .block-container{
        padding-top: 1rem;
        padding-bottom: 0rem;
    }
    /* Style the sidebar */
    .sidebar .sidebar-content {
        background-color: #f0f2f6; /* Light gray background */
    }
    /* Style filter elements */
    .stSlider, .stSelectbox, .stMultiselect, .stTextInput, .stDateInput {
        margin-bottom: 0.5rem;
    }
    /* Style the dataframe */
    .dataframe {
        border-radius: 0.5rem;
        overflow: hidden; /* For rounded corners */
    }
    /* Style expander */
    .stExpander {
        border: 1px solid #d9d9d9;
        border-radius: 0.5rem;
    }
</style>
""",
    unsafe_allow_html=True,
)

# --- Utility Functions ---

@st.cache_data(show_spinner=False)  # Changed to st.cache_data
def load_excel(file, sheet_name=None, header_row=0):
    """Loads an Excel file, handling sheet selection, header row, and errors.

    Args:
        file (BytesIO or str): The Excel file (either a BytesIO object or a file path).
        sheet_name (str, optional): The name of the sheet to load. Defaults to None (load sheet names).
        header_row (int, optional): The row number (0-indexed) to use as the header. Defaults to 0.

    Returns:
        tuple: A tuple containing:
            - pd.DataFrame: The loaded DataFrame (or None if only sheet names are loaded).
            - list: A list of sheet names (or None if a specific sheet is loaded).
            - int: The header row number used.
    """
    try:
        if isinstance(file, BytesIO):
            if sheet_name is None:
                excel_file = openpyxl.load_workbook(file, read_only=True)
                sheet_names = excel_file.sheetnames
                return None, sheet_names, None  # Return None, sheet_names, header_row
            df = pd.read_excel(file, engine='openpyxl', sheet_name=sheet_name, header=header_row)
            return df, None, header_row  # Return df, None, header_row
        # Assuming file is a file path (for testing)
        if sheet_name is None:
            excel_file = openpyxl.load_workbook(file, read_only=True)
            sheet_names = excel_file.sheetnames
            return None, sheet_names, None
        df = pd.read_excel(file, engine='openpyxl', sheet_name=sheet_name, header=header_row)
        return df, None, header_row
    except Exception as e:
        st.error(f"Error loading Excel file or sheet: {e}")
        return pd.DataFrame(), None, None

@st.cache_data(show_spinner=False)
def _build_query(column_filters, logic):
    """Builds a pandas query string for efficient filtering.

    Args:
        column_filters (dict): A dictionary of column filters.
        logic (str): The logical operator ("AND" or "OR").

    Returns:
        str: The pandas query string.
    """
    query_parts = []
    for column, values in column_filters.items():
        if not values:
            continue

        # Handle different data types
        if isinstance(values[0], (int, float, np.number)):
            # Numeric range (already handled as a list of two values)
            if len(values) == 2:
                min_val, max_val = values
                query_parts.append(f"`{column}` >= {min_val} and `{column}` <= {max_val}")
        elif isinstance(values[0], (pd.Timestamp, type(pd.NaT))):
            # Date range (already handled as a list of two values)
            if len(values) == 2:
                start_date, end_date = values
                # Convert to string representation for query
                start_str = start_date.strftime('%Y-%m-%d')
                end_str = end_date.strftime('%Y-%m-%d')
                query_parts.append(f"`{column}` >= '{start_str}' and `{column}` <= '{end_str}'")
        else:  # String or other types
            # Use .isin() for string lists (vectorized and efficient)
            values_str = ', '.join([f"'{v}'" for v in values])  # Quote string values
            query_parts.append(f"`{column}` in [{values_str}]")

    if logic == "AND":
        return " and ".join(query_parts)
    elif logic == "OR":
        return " or ".join(query_parts)
    return ""  # Return empty string if no filters

@st.cache_data(show_spinner=False)
def _apply_search(df, search_term, use_regex=False):
    """Applies a search term to the DataFrame.

    Args:
        df (pd.DataFrame): The DataFrame to search.
        search_term (str): The search term.
        use_regex (bool): Whether to use regular expressions.

    Returns:
        pd.DataFrame: The filtered DataFrame.
    """
    if not search_term:
        return df

    if use_regex:
        try:
            regex = re.compile(search_term, re.IGNORECASE)
        except re.error as e:
            st.error(f"Invalid regex pattern: {e}")
            return df
        # Corrected: Use str.contains directly with the regex object
        mask = df.apply(lambda row: row.astype(str).str.contains(regex, na=False).any(), axis=1)
        df = df[mask]  # Apply the mask *once*
    else:
        # Lowercase the search term for case-insensitive search
        search_term = search_term.lower()

        if '&' in search_term:
            terms = [term.strip() for term in search_term.split('&')]
            for term in terms:
                # Corrected: Use .any(axis=1) for row-wise OR
                mask = df.apply(lambda row: row.astype(str).str.lower().str.contains(term, na=False).any(), axis=1)
                df = df[mask]  # Apply mask in each iteration
        elif '|' in search_term:
            terms = [term.strip() for term in search_term.split('|')]
            # Corrected: Use a single mask with str.contains and join terms with |
            mask = df.apply(lambda row: row.astype(str).str.lower().str.contains('|'.join(terms), na=False).any(), axis=1)
            df = df[mask]
        elif '!' in search_term:
            term = search_term.replace('!', '').strip()  # Corrected: Remove !
            # Corrected: Use .any(axis=1) and ~ for NOT
            mask = df.apply(lambda row: ~row.astype(str).str.lower().str.contains(term, na=False).any(), axis=1)
            df = df[mask]
        else:
            # Corrected: Use .any(axis=1)
            mask = df.apply(lambda row: row.astype(str).str.lower().str.contains(search_term, na=False).any(), axis=1)
            df = df[mask]

    return df

@st.cache_data(show_spinner=False)
def filter_dataframe(df, column_filters, search_term, logic, use_regex=False):
    """Filters DataFrame using pandas query and vectorized search.

    Args:
        df (pd.DataFrame): The DataFrame to filter.
        column_filters (dict): A dictionary of column filters.
        search_term (str): The search term.
        logic (str): The logical operator ("AND" or "OR").
        use_regex (bool): Whether to use regular expressions for the search.

    Returns:
        pd.DataFrame: The filtered DataFrame.
    """
    query_string = _build_query(column_filters, logic)
    if query_string:
        try:
            df_filtered = df.query(query_string)
        except Exception as e:
            st.error(f"Error in query: {e}")
            return df  # Return original DataFrame on error
    else:
        df_filtered = df.copy()

    df_filtered = _apply_search(df_filtered, search_term, use_regex)
    return df_filtered

def download_excel(df, filename="filtered_data.xlsx"):
    """Generates a download link for the filtered DataFrame as an Excel file."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        # No need to explicitly save: writer.__exit__ handles it
    excel_data = output.getvalue()
    st.download_button(
        label="Download Filtered Data as Excel",
        data=excel_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def download_csv(df, filename="filtered_data.csv"):
    """Generates a download link for the filtered DataFrame as a CSV file."""
    output = BytesIO()
    df.to_csv(output, index=False)
    csv_data = output.getvalue()
    st.download_button(
        label="Download Filtered Data as CSV",
        data=csv_data,
        file_name=filename,
        mime="text/csv",
    )

def download_json(df, filename="filtered_data.json"):
    """Generates a download link for the filtered DataFrame as a JSON file."""
    output = BytesIO()
    df.to_json(output, orient='records', lines=True)
    json_data = output.getvalue()
    st.download_button(
        label="Download Filtered Data as JSON",
        data=json_data,
        file_name=filename,
        mime="application/json",
    )

def get_summary_statistics(df):
    """Calculates summary statistics, handling datetime columns correctly."""
    if df.empty:
        return "No data to display."

    # Separate numeric, datetime, and other columns
    numeric_cols = df.select_dtypes(include=np.number).columns
    datetime_cols = df.select_dtypes(include='datetime').columns
    other_cols = df.select_dtypes(exclude=[np.number, 'datetime']).columns

    # Calculate statistics separately and combine
    summary = {}
    if not numeric_cols.empty:
        summary['Numeric Columns'] = df[numeric_cols].describe()
    if not datetime_cols.empty:
        summary['Datetime Columns'] = df[datetime_cols].describe(datetime_is_numeric=False)
    if not other_cols.empty:
        summary['Other Columns'] = df[other_cols].describe()

    # Display each summary DataFrame
    for title, stats_df in summary.items():
        st.write(f"**{title}:**")
        st.dataframe(stats_df)

def aggregate_data(df):
    """Allows users to perform basic aggregation operations."""
    if df.empty:
        st.warning("No data to aggregate.")
        return

    st.sidebar.title("Data Aggregation")
    agg_columns = st.sidebar.multiselect("Select columns to aggregate", options=df.columns)
    agg_functions = st.sidebar.multiselect("Select aggregation functions", options=['sum', 'mean', 'count', 'max', 'min', 'std'])

    if agg_columns and agg_functions:  # Corrected variable name
        try:
            agg_df = df[agg_columns].agg(agg_functions)
            st.write("### Aggregated Data")
            st.dataframe(agg_df)
        except Exception as e:
            st.error(f"Error in aggregation: {e}")

def create_pivot_table(df):
    """Allows users to create pivot tables."""
    if df.empty:
        st.warning("No data to create pivot table.")
        return

    st.sidebar.title("Pivot Table")
    index_cols = st.sidebar.multiselect("Select index columns", options=df.columns)
    values_cols = st.sidebar.multiselect("Select value columns", options=df.columns)
    agg_func = st.sidebar.selectbox("Select aggregation function", options=['sum', 'mean', 'count', 'max', 'min', 'std'])

    if index_cols and values_cols:
        try:
            pivot_df = pd.pivot_table(df, values=values_cols, index=index_cols, aggfunc=agg_func)
            st.write("### Pivot Table")
            st.dataframe(pivot_df)
        except Exception as e:
            st.error(f"Error in pivot table creation: {e}")

def clean_data(df):
    """Performs data cleaning operations and returns a NEW DataFrame."""
    if df.empty:
        st.warning("No data to clean.")
        return df  # Return the original empty DataFrame

    df_cleaned = df.copy()  # Create a copy!

    st.sidebar.title("Data Cleaning")
    remove_duplicates = st.sidebar.checkbox("Remove Duplicates")
    drop_na = st.sidebar.checkbox("Drop Missing Values")
    rename_columns = st.sidebar.checkbox("Rename Columns")  # Corrected variable name

    if remove_duplicates:
        df_cleaned = df_cleaned.drop_duplicates()

    if drop_na:
        df_cleaned = df_cleaned.dropna()

    if rename_columns:  # Corrected variable name
        st.write("### Rename Columns")
        new_column_names = st.text_input("Enter new column names (comma-separated)", value=','.join(df_cleaned.columns))
        new_column_names = [name.strip() for name in new_column_names.split(',')]
        if len(new_column_names) == len(df_cleaned.columns):
            df_cleaned.columns = new_column_names
        else:
            st.error("Number of new column names must match the number of existing columns.")

    return df_cleaned  # Return the modified copy

def advanced_filtering(df):
    """Allows users to perform advanced filtering and returns a NEW DataFrame."""
    if df.empty:
        st.warning("No data to filter.")
        return df

    df_filtered = df.copy() # Create a copy

    st.sidebar.title("Advanced Filtering")
    filter_by_condition = st.sidebar.checkbox("Filter by Condition")

    if filter_by_condition:
        condition = st.sidebar.text_input("Enter filtering condition (e.g., `column_name > value`)")
        if condition:
            try:
                df_filtered = df_filtered.query(condition)
            except Exception as e:
                st.error(f"Error in advanced filtering: {e}")
    return df_filtered # Return the filtered copy

# --- Main App Logic ---

st.title("Sheet Sherpa")
st.write("Upload an Excel/CSV file to analyze its contents with advanced filtering and search.")

# Sidebar for file upload
st.sidebar.title("File Selector")
uploaded_file = st.sidebar.file_uploader("Upload an Excel file", type=["xlsx", "csv"])

# Button to clear filters (resets the session state)
if st.sidebar.button("Clear Filters"):  # Corrected: String, not a function call
    st.session_state.clear()
    st.rerun()

# Initialize df and sheet_names here
df = pd.DataFrame()
sheet_names = None
selected_sheet = None  # Initialize selected_sheet
header_row = 0 # Initialize header_row

if uploaded_file:
    file_content = BytesIO(uploaded_file.read())

    # --- Sheet and Header Selection ---
    df, sheet_names, _ = load_excel(file_content)  # Initial load


    if sheet_names:
        selected_sheet = st.sidebar.selectbox("Select a sheet", sheet_names)  # Corrected: String

    # Header Row Selection (placed inside sidebar)
    header_row = st.sidebar.number_input("Header Row (0-indexed)", min_value=0, value=0, step=1)  # Corrected: String

    if selected_sheet:
        df, _, header_row = load_excel(file_content, sheet_name=selected_sheet, header_row=header_row)
    elif sheet_names is None and df is None:
        st.warning("No sheets found in the uploaded file.")
        # df already initialized as empty DataFrame
    elif df is not None: #Single sheet excel file.
        df, _, header_row = load_excel(file_content, header_row=header_row) #Load with header.

    if not df.empty:
        # --- Data Cleaning ---
        df = clean_data(df)

        # --- Advanced Filtering ---
        df = advanced_filtering(df)

        # --- Sidebar Filters ---
        st.sidebar.title("Filters")  # Corrected: String

        # Dynamic Column Filters
        column_filters = {}  # Corrected variable name
        with st.sidebar.expander("Column Filters", expanded=True):  # Corrected: String
            for column in df.columns:
                if pd.api.types.is_numeric_dtype(df[column]):
                    min_val = df[column].min()
                    max_val = df[column].max()
                    if pd.isna(min_val) or pd.isna(max_val):
                        selected_values = []
                        st.write(f"No Filter for {column} (all values are NaN)")
                    else:
                        selected_range = st.slider(  # Corrected: No f-string
                            f"Range for {column}",
                            min_value=float(min_val),
                            max_value=float(max_val),
                            value=[float(min_val), float(max_val)],
                            step=(0.01 if (max_val - min_val) < 1 else (1.0 if (max_val - min_val) < 1000 else 100.0))
                        )
                        selected_values = list(selected_range)
                elif pd.api.types.is_datetime64_any_dtype(df[column]):  # Corrected: No f-string
                    min_date = df[column].min()
                    max_date = df[column].max()
                    if pd.isna(min_date) or pd.isna(max_date):
                        selected_values = []
                        st.write(f"No Filter for {column} (all values are NaT)")
                    else:
                        selected_dates = st.date_input(  # Corrected: No f-string
                            f"Date range for {column}",
                            value=(min_date, max_date),
                            min_value=min_date,
                            max_value=max_date,
                        )
                        if len(selected_dates) == 2:
                            selected_values = [pd.Timestamp(d) for d in selected_dates]
                        else:
                            selected_values = []
                else:
                    unique_values = df[column].dropna().unique().tolist()
                    unique_values = sorted(unique_values, key=lambda x: (isinstance(x, str), x))
                    selected_values = st.multiselect(f"Select values for {column}", options=unique_values)  # Corrected: No f-string

                column_filters[column] = selected_values

        # Search Term Input
        search_keyword = st.sidebar.text_input("Enter keyword to search across all columns")  # Corrected: String
        use_regex = st.sidebar.checkbox("Use Regex for Search")  # Corrected: String
        with st.sidebar.expander("Search Operators"):  # Corrected: String
            st.markdown("- Use `&` for AND (e.g., `word1 & word2`)")
            st.markdown("- Use `|` for OR (e.g., `word1 | word2`)")
            st.markdown("- Use `!` for NOT (e.g., `!word`)")

        # Logic selection for multi-filters
        logic = st.sidebar.radio("Filter Logic for Column Filters", ["AND", "OR"], index=0)  # Corrected: String

        # --- Filter and Display Data ---
        with st.spinner("Filtering data..."):  # Corrected: String
            start_time = time.time()
            df_filtered = filter_dataframe(df, column_filters, search_keyword, logic, use_regex)
            end_time = time.time()
            st.write(f"Data filtered in {end_time - start_time:.2f} seconds")

        # --- Main Content Area ---
        st.dataframe(df_filtered, height=500, use_container_width=True)  # Corrected: st.dataframe

        # Download Buttons
        if selected_sheet:
            download_excel(df_filtered, filename=f"filtered_data_{selected_sheet}_header_{header_row}.xlsx")
            download_csv(df_filtered, filename=f"filtered_data_{selected_sheet}_header_{header_row}.csv")
            download_json(df_filtered, filename=f"filtered_data_{selected_sheet}_header_{header_row}.json")
        else:
            download_excel(df_filtered, filename=f"filtered_data_header_{header_row}.xlsx")
            download_csv(df_filtered, filename=f"filtered_data_header_{header_row}.csv")
            download_json(df_filtered, filename=f"filtered_data_header_{header_row}.json")

        # Display selected row details
        if not df_filtered.empty:
            with st.expander("View Selected Row Details"):  # Corrected: String
                selected_row_index = st.selectbox("Select a row to view details", df_filtered.index)  # Corrected: String
                if selected_row_index is not None:
                    row_data = df_filtered.loc[selected_row_index]
                    for col in df_filtered.columns:
                        st.write(f"**{col}:** {row_data[col]}")

        # --- Summary Statistics ---
        with st.expander("Summary Statistics"):  # Corrected: String
            get_summary_statistics(df_filtered)

        # --- Data Aggregation ---
        aggregate_data(df_filtered)

        # --- Pivot Tables ---
        create_pivot_table(df_filtered)

        # --- Data Visualization ---
        st.sidebar.title("Data Visualization")  # Corrected: String
        if not df_filtered.empty:
            numeric_cols = df_filtered.select_dtypes(include=np.number).columns
            if not numeric_cols.empty:
                x_axis = st.sidebar.selectbox("Select X-axis", options=df_filtered.columns)  # Corrected: String
                y_axis = st.sidebar.selectbox("Select Y-axis", options=numeric_cols)  # Corrected: String
                chart_type = st.sidebar.radio("Select Chart Type", ["Bar Chart", "Line Chart", "Histogram", "Scatter Plot", "Box Plot", "Heatmap", "Time Series"], index=0) # Corrected: String

                if x_axis and y_axis:
                    if chart_type == "Bar Chart":
                        chart = alt.Chart(df_filtered).mark_bar().encode(
                            x=alt.X(f"{x_axis}:N", title=x_axis),
                            y=alt.Y(f"{y_axis}:Q", title=y_axis)
                        ).properties(
                            width=600,
                            height=400
                        )
                    elif chart_type == "Line Chart":
                        chart = alt.Chart(df_filtered).mark_line().encode(
                            x=alt.X(f"{x_axis}:N", title=x_axis),
                            y=alt.Y(f"{y_axis}:Q", title=y_axis)
                        ).properties(
                            width=600,
                            height=400
                        )
                    elif chart_type == "Histogram":
                        chart = alt.Chart(df_filtered).mark_bar().encode(
                            x=alt.X(f"{x_axis}:Q", title=x_axis, bin=True),
                            y=alt.Y("count()", title="Count")
                        ).properties(
                            width=600,
                            height=400
                        )
                    elif chart_type == "Scatter Plot":
                        chart = alt.Chart(df_filtered).mark_circle().encode(
                            x=alt.X(f"{x_axis}:Q", title=x_axis),
                            y=alt.Y(f"{y_axis}:Q", title=y_axis)
                        ).properties(
                            width=600,
                            height=400
                        )
                    elif chart_type == "Box Plot":
                        chart = alt.Chart(df_filtered).mark_boxplot().encode(
                            x=alt.X(f"{x_axis}:N", title=x_axis),
                            y=alt.Y(f"{y_axis}:Q", title=y_axis)
                        ).properties(
                            width=600,
                            height=400
                        )
                    elif chart_type == "Heatmap":
                        # Corrected Heatmap implementation (from previous response)
                        value_col = st.sidebar.selectbox("Select Value for Heatmap", options=numeric_cols)
                        if value_col:  # Make sure a value column is selected
                            heatmap_df = df_filtered.pivot_table(index=x_axis, columns=y_axis, values=value_col, aggfunc='mean')
                            # Reset index for Altair
                            heatmap_df = heatmap_df.reset_index().melt(id_vars=x_axis, var_name=y_axis, value_name='_heatmap_value')

                            chart = alt.Chart(heatmap_df).mark_rect().encode(
                                x=alt.X(f"{x_axis}:O", title=x_axis),  # Use :O for Ordinal
                                y=alt.Y(f"{y_axis}:O", title=y_axis),  # Use :O for Ordinal
                                color=alt.Color('_heatmap_value:Q', title="Mean Value")  # Use the calculated value
                            ).properties(
                                width=600,
                                height=400
                            )
                        else:
                            chart = None
                            st.sidebar.warning("Please select a value column for the heatmap.")

                    elif chart_type == "Time Series":
                        if pd.api.types.is_datetime64_any_dtype(df_filtered[x_axis]):
                            chart = alt.Chart(df_filtered).mark_line().encode(
                                x=alt.X(f"{x_axis}:T", title=x_axis),
                                y=alt.Y(f"{y_axis}:Q", title=y_axis)
                            ).properties(
                                width=600,
                                height=400
                            )
                        else:
                            st.error(f"X-axis column {x_axis} is not a datetime column.")
                            chart = None

                    if chart:
                        st.altair_chart(chart, use_container_width=True)  # Use container width
            else:
                st.warning("No numeric columns available for visualization.")

else:
    if sheet_names is None:
        st.warning("The uploaded file is empty or could not be loaded.")
